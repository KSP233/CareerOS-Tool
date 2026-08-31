"""CareerOS clean-room job, document, and manual-application services."""
from __future__ import annotations

import csv, difflib, hashlib, html, json, logging, re, shutil, sys, webbrowser, zipfile
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path

from app.ai_manager import AIManager
from app.database import Database
from app.pdf_export import render_cv_document_pdf, render_cv_letter_pdf, render_resume_document_pdf, render_resume_pdf, safe_filename
from app.docx_export import export_docx_pdf, render_cv_letter_docx, render_professional_one_page, render_structured_cv_docx, render_structured_resume_docx
from app.resume_models import ResumeDocument, ResumeLayout, ResumeStyle, resume_document_from_dict, resume_document_to_dict, resume_layout_from_dict, resume_layout_to_dict, resume_style_from_dict, resume_style_to_dict
from app.resume_ingestion import blocks_to_dict, parse_docx_resume
from app.resume_parser import parse_resume_text, resume_document_to_text
from app.resume_integrity import ResumeIntegrityValidator, apply_resume_operations, apply_resume_operations_best_effort
from app.prompts import CV_GENERATION_PROMPT, JOB_ANALYSIS_PROMPT, RESUME_OPTIMIZATION_PROMPT, TRANSLATION_PROMPT
from app.validators import recommendation, validate_job_analysis, validate_resume_changes
from config import SUPPORTED_RESUME_IMPORT_EXTENSIONS, get_paths, load_settings

log = logging.getLogger(__name__)
_PAY = re.compile(r"(?P<mark>CA\$|C\$|CAD\s*\$?|\$)\s*(?P<lo>\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)\s*(?:-|–|—|to)\s*(?:(?:CA\$|C\$|CAD\s*\$?|\$)\s*)?(?P<hi>\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)", re.I)

def generation_instruction(kind: str) -> str:
    """Return the shared user instruction for Resume/CV generation.

    `resume` remains a migration fallback for settings written by older builds.
    Other AI actions intentionally do not inherit resume/CV content instructions.
    """
    if str(kind).casefold() not in {"resume", "cv"}:
        return ""
    prompts = load_settings().get("generation_prompts", {})
    note = str(prompts.get("general", prompts.get("resume", ""))).strip()
    return f"\n\nUSER-APPROVED EXTRA INSTRUCTIONS:\n{note}" if note else ""

def weighted_match_score(rule_score: int, ai_score: int, settings: dict | None = None) -> int:
    value = (settings or load_settings()).get("match_weights", {}).get("rule", 70)
    try: rule = min(100, max(0, int(value)))
    except (ValueError, TypeError): rule = 70
    return round((rule_score * rule + ai_score * (100 - rule)) / 100)

class _JobHtmlText(HTMLParser):
    def __init__(self): super().__init__(convert_charrefs=True); self.parts=[]
    def handle_starttag(self, tag, attrs):
        if tag in {"p","div","li","br","h1","h2","h3","h4","h5","h6","tr"}: self.parts.append("\n")
        if tag == "li": self.parts.append("• ")
    def handle_endtag(self, tag):
        if tag in {"p","div","li","h1","h2","h3","h4","h5","h6","tr"}: self.parts.append("\n")
    def handle_data(self, data): self.parts.append(data)


def compact_job_text(text: str) -> str:
    """Normalize copied or scraped posting text without changing its facts."""
    blocks, last_blank = [], False; source = str(text or "")[:2_000_000]
    if re.search(r"</?[A-Za-z][^>]*>", source):
        parser = _JobHtmlText()
        try: parser.feed(source); parser.close(); source = "".join(parser.parts)
        except Exception: source = re.sub(r"<[^>]{0,1000}>", " ", source)
    source = html.unescape(source).replace("\r", "").replace("\\-", "-").replace("\\'", "'")
    for raw in source.split("\n"):
        line = re.sub(r"\s+", " ", raw).strip(); line = re.sub(r"^#{1,6}\s*", "", line); line = re.sub(r"\*{1,2}([^*]+)\*{1,2}", r"\1", line)
        if not line:
            if blocks and not last_blank: blocks.append("")
            last_blank = True; continue
        if re.match(r"^[*+•-]\s+", line): line = "• " + re.sub(r"^[*+•-]\s+", "", line)
        blocks.append(line); last_blank = False
    return "\n".join(blocks).strip()

def extract_job_facts(title: str, description: str) -> dict[str, object]:
    """Extract only plainly stated dates and requirements for display; do not infer them."""
    raw = f"{title}\n{description}"; lines=[]
    for part in re.split(r"[\r\n]+", raw):
        item = re.sub(r"[*_`#]+", "", part).strip(" -:;\t")
        if 4 <= len(item) <= 420 and item not in lines: lines.append(item)
    season = re.search(r"\b(Spring|Summer|Fall|Autumn|Winter)\s+(20\d{2})\b", title, re.I)
    dated = re.search(r"\b(?:start(?:ing)? date|expected start|commenc(?:ement|ing)|begin(?:ning)?)\s*(?:is|:|-)?\s*((?:January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+\d{1,2}(?:st|nd|rd|th)?)?(?:,?\s+20\d{2})?|(?:Spring|Summer|Fall|Autumn|Winter)\s+20\d{2}|20\d{2}-\d{2}-\d{2})", raw, re.I)
    start = f"{season.group(1).title()} {season.group(2)}" if season else (dated.group(1).strip() if dated else "Not stated")
    education = re.compile(r"\b(bachelor(?:'|’)?s?|master(?:'|’)?s?|ph\.?d\.?|degree|diploma|currently enrolled|engineering technology program)\b", re.I)
    other = re.compile(r"\b(?:\d+\s*(?:\+|[-–]\s*\d+)?\s+years?|must|required|eligible|clearance|citizen|work authorization|driver(?:'|’)??s? licen[cs]e|travel|on-?site|shift|bilingual|french|professional engineer|p\.eng\.|oiq|communication)\b", re.I)
    def unique(values, limit):
        seen=set(); result=[]
        for value in values:
            key=value.casefold()
            if key not in seen: result.append(value); seen.add(key)
            if len(result) == limit: break
        return result
    edu = unique([line for line in lines if education.search(line) and "privacy" not in line.casefold()], 4)
    return {"start_date":start, "education":edu, "other_requirements":unique([line for line in lines if other.search(line) and line not in edu],14)}

def evaluate_requirement(requirement: str, candidate_facts: str) -> str:
    req, facts = requirement.casefold(), candidate_facts.casefold()
    denials=(("driver",("no driver's license","do not have a driver's license")),("travel",("cannot travel","not willing to travel")),("authorization",("not authorized","require sponsorship")),("onsite",("remote only","cannot work onsite")))
    if any(flag in req and any(no in facts for no in nos) for flag,nos in denials): return "DOES_NOT_MEET"
    years=re.search(r"\b(?:minimum of\s+)?(\d+)\s*(?:\+|[-–]\s*\d+|to\s+\d+)?\s+years?",req)
    if years:
        known=[int(v) for v in re.findall(r"\b(\d+)\+?\s+years?",facts)]; return "LIKELY_MET" if known and max(known)>=int(years.group(1)) else "NOT_CONFIRMED"
    groups=[]
    if "bachelor" in req: groups.append(("bachelor","b.eng","basc"))
    if "master" in req: groups.append(("master","m.eng","msc"))
    disciplines=tuple(x for x in ("mechanical","aerospace","civil","electrical","mechatronics","software","computer") if x in req)
    if disciplines: groups.append(disciplines)
    if "french" in req: groups.append(("french",))
    if "work authorization" in req: groups.append(("work authorization","authorized to work"))
    return "LIKELY_MET" if groups and all(any(word in facts for word in group) for group in groups) else "NOT_CONFIRMED"

class JobService:
    def __init__(self, db: Database, ai: AIManager):
        self.db,self.ai=db,ai; self._refresh_derived_fields()

    @staticmethod
    def _value(value) -> str:
        if value is None: return ""
        try:
            import pandas as pd
            if pd.isna(value): return ""
        except (ImportError,TypeError,ValueError): pass
        return value.isoformat() if hasattr(value,"isoformat") else str(value)

    @staticmethod
    def _description_hash(text: str) -> str: return hashlib.sha256(text.encode("utf-8","ignore")).hexdigest()

    @classmethod
    def _salary_from_description(cls, description: str) -> str:
        values=[]
        for match in _PAY.finditer(str(description or "")):
            try: lo,hi=float(match["lo"].replace(",","")),float(match["hi"].replace(",",""))
            except ValueError: continue
            if hi >= lo: values.append((lo,hi,match["mark"].upper().replace(" ","")))
        if not values: return ""
        money=lambda n: f"{n:,.2f}".rstrip("0").rstrip(".")
        sign="CA$" if any(mark.startswith("CA") or mark=="C$" for _,_,mark in values) else "$"
        return f"{sign}{money(min(x[0] for x in values))} - {sign}{money(max(x[1] for x in values))}" + (" (varies by location)" if len(values)>1 else "")

    @classmethod
    def _salary(cls, row, description: str) -> str:
        low,high,unit,currency=(cls._value(row.get(k)) for k in ("min_amount","max_amount","interval","currency"))
        stated=" ".join(x for x in (" - ".join(x for x in (low,high) if x),currency,unit) if x)
        return stated or cls._salary_from_description(description)

    def _refresh_derived_fields(self) -> None:
        # Search keywords describe desired jobs, not candidate evidence.  Older
        # builds used them as a fallback and displayed ~50% scores even before
        # a resume had been imported.
        context=self._resume_text()
        for job in self.db.jobs():
            changes={}; description=str(job.get("description") or "")
            if not str(job.get("salary") or "").strip() and description: changes["salary"]=self._salary_from_description(description)
            if not str(job.get("start_date") or "").strip() and description: changes["start_date"]=extract_job_facts(job.get("title",""),description)["start_date"]
            if not context:
                if any(job.get(key) is not None for key in ("rule_score", "ai_score", "match_score")):
                    changes.update(rule_score=None, ai_score=None, match_score=None, match_reason="", strengths_json="[]", missing_json="[]", risks_json="[]", requirements_json="[]", preferred_json="[]", recommendation="", model_used=None)
            elif job.get("rule_score") is None and description:
                changes["rule_score"] = self._rule_score(context, job.get("title", "") + "\n" + description, job.get("location", ""))[0]
            if changes: self.db.update_job(job["id"],**changes)

    @staticmethod
    def _resume_text() -> str:
        paths=get_paths(); copies=sorted(paths.resumes_original.glob("original-*.txt"),key=lambda p:p.stat().st_mtime); configured=str(load_settings().get("resume_path","")).strip(); path=copies[-1] if copies else (Path(configured) if configured else None)
        return path.read_text(encoding="utf-8",errors="replace") if path and path.is_file() and path.suffix.casefold() in {".txt",".md"} else ""

    @staticmethod
    def _rule_score(candidate: str, posting: str, location: str) -> tuple[int,list[str],list[str]]:
        known=set(re.findall(r"[a-z][a-z0-9+#.-]{2,}",candidate.casefold())); requested=list(dict.fromkeys(re.findall(r"[A-Za-z][A-Za-z0-9+#.-]{2,}",posting)))
        matches=[word for word in requested if word.casefold() in known][:20]; keywords={"engineering","mechanical","aerospace","design","python","matlab","simulink","ansys","catia","fusion","cfd","fea","testing","cad"}; missing=[word for word in requested if word.casefold() in keywords and word.casefold() not in known]
        skill=min(65,20+3*len({word.casefold() for word in matches})); senior=25 if re.search(r"\b(senior|lead|principal|manager|[5-9]\+? years|10\+? years)\b",posting,re.I) else 0; location_bonus=5 if any(city in str(location).casefold() for city in ("ottawa","montreal","montréal","toronto","quebec","québec","remote")) else 0
        return max(0,min(100,skill+15+location_bonus-senior)),matches,missing

    def _store(self, raw: dict, context: str) -> bool:
        description=compact_job_text(self._value(raw.get("description"))); title=self._value(raw.get("title")); job={"company":self._value(raw.get("company")),"title":title,"location":self._value(raw.get("location")),"salary":self._value(raw.get("salary")) or self._salary(raw,description),"source":self._value(raw.get("site") or raw.get("source")),"url":self._value(raw.get("job_url_direct") or raw.get("job_url") or raw.get("url")),"description":description,"date_posted":self._value(raw.get("date_posted")),"description_hash":self._description_hash(description),"start_date":extract_job_facts(title,description)["start_date"]}
        job_id,added=self.db.upsert_job(job)
        if context and description: self.db.update_job(job_id,rule_score=self._rule_score(context,title+"\n"+description,job["location"])[0])
        return added

    def search(self, progress=lambda message: None, cancelled=lambda: False) -> dict:
        # JobSpy 1.1.82 requires a vulnerable markdownify version even when its
        # caller requests HTML.  Install CareerOS's bounded compatibility shim
        # before importing JobSpy; the unsafe distribution is excluded from the
        # portable executable by CareerOS.spec.
        from app import safe_markdownify
        sys.modules["markdownify"] = safe_markdownify
        try: from jobspy import scrape_jobs
        except ImportError as exc: raise RuntimeError("python-jobspy is not installed") from exc
        settings=load_settings()["search"]; context=self._resume_text(); added=existing=errors=index=0; total=len(settings["locations"])*len(settings["queries"])
        for location in settings["locations"]:
            for query in settings["queries"]:
                if cancelled():
                    return {"added":added,"existing":existing,"errors":errors,"cancelled":True}
                index+=1; progress(f"Searching {index}/{total}: {query} in {location}")
                try:
                    # JobSpy 1.1.82 pins a vulnerable markdownify release.
                    # Request HTML so that converter is never called; CareerOS
                    # performs bounded stdlib HTML-to-text conversion in _store.
                    frame=scrape_jobs(site_name=settings["sites"],search_term=query,location=location,results_wanted=settings["results_per_search"],hours_old=settings["hours_old"],country_indeed="canada",distance=settings["distance"],linkedin_fetch_description=True,description_format="html")
                    for _,row in frame.iterrows():
                        if self._store(row.to_dict(),context): added+=1
                        else: existing+=1
                except Exception as exc: errors+=1; log.warning("Search failed: %s",exc); progress(f"Search warning: {query} / {location}: {exc}")
        return {"added":added,"existing":existing,"errors":errors,"cancelled":False}

    def import_file(self, source: str) -> dict:
        path=Path(source)
        if path.suffix.casefold()==".csv":
            with path.open("r",encoding="utf-8-sig",newline="") as handle: records=list(csv.DictReader(handle))
        elif path.suffix.casefold()==".json":
            records=json.loads(path.read_text(encoding="utf-8")); records=records if isinstance(records,list) else [records]
        else: raise ValueError("Import a .csv or .json job file")
        context=self._resume_text(); added=existing=0
        for record in records:
            if isinstance(record,dict) and self._value(record.get("title")):
                record={**record,"source":record.get("source") or "import"}; is_new=self._store(record,context); added += int(is_new); existing += int(not is_new)
        return {"added":added,"existing":existing}

    def add_manual(self, values: dict) -> int:
        description=str(values.get("description") or ""); record={**values,"salary":values.get("salary") or self._salary_from_description(description),"description_hash":self._description_hash(description),"start_date":extract_job_facts(str(values.get("title") or ""),description)["start_date"]}; job_id,_=self.db.upsert_job(record); context=self._resume_text()
        if context and description: self.db.update_job(job_id,rule_score=self._rule_score(context,str(record.get("title"))+"\n"+description,str(record.get("location","")))[0])
        return job_id

    def analyze(self, job_id: int, resume: str, selected_model: str | None = None, *, progress=None, cancelled=None) -> dict:
        job=self.db.job(job_id)
        if not job: raise ValueError("Job not found")
        if not str(resume or "").strip(): raise ValueError("Import an original DOCX resume before analyzing job matches.")
        rule,matched,missing=self._rule_score(resume,job["title"]+"\n"+job["description"],job["location"])
        resume_payload = resume.removeprefix("RESUME:\n")
        resume_excerpt, job_excerpt = resume_payload[:12000], str(job["description"] or "")[:12000]
        payload=f"RESUME:\n{resume_excerpt}\n\nJOB:\nTitle: {job['title']}\nCompany: {job['company']}\nLocation: {job['location']}\n{job_excerpt}"
        log.debug("job_analysis_payload job_id=%s model=%s resume_chars=%s job_description_chars=%s payload_chars=%s", job_id, selected_model or "routed", len(resume_excerpt), len(job_excerpt), len(payload))
        analysis,model=self.ai.generate_json("job_analysis",JOB_ANALYSIS_PROMPT,payload,selected_model,progress=progress,cancelled=cancelled); analysis=validate_job_analysis(analysis); final=weighted_match_score(rule,analysis["ai_score"]); analysis.update({"match_score":final,"rule_score":rule,"recommendation":recommendation(final),"required_matches":list(dict.fromkeys(matched+analysis["required_matches"])),"missing_skills":list(dict.fromkeys(missing+analysis["missing_skills"])),"model_used":model})
        self.db.update_job(job_id,rule_score=rule,ai_score=analysis["ai_score"],match_score=final,match_reason=analysis["reason"],strengths_json=json.dumps(analysis["strengths"]),missing_json=json.dumps(analysis["missing_skills"]),risks_json=json.dumps(analysis["risks"]),requirements_json=json.dumps(analysis["required_matches"]),preferred_json=json.dumps(analysis["preferred_matches"]),recommendation=analysis["recommendation"],model_used=model); return analysis

    def translate(self, job_id: int, progress=lambda message: None, cancelled=lambda: False) -> dict:
        job=self.db.job(job_id)
        if not job: raise ValueError("Job not found")
        text=compact_job_text(job["description"]); digest=self._description_hash(text)
        if job.get("translation_zh") and job.get("translation_hash")==digest: return {"translation":job["translation_zh"],"model_used":job.get("translation_model") or "cached","cached":True}
        if not text: raise ValueError("Job description is empty")
        pieces=[text[i:i+6000] for i in range(0,len(text),6000)]; result=[]; models=[]; selected_model=load_settings().get("ai_mode", "Auto")
        for index,piece in enumerate(pieces,1):
            if cancelled():
                return {"translation":"\n\n".join(result),"model_used":", ".join(dict.fromkeys(models)) or "cancelled","cached":False,"cancelled":True}
            progress(f"Translating {index}/{len(pieces)} with {selected_model}..."); response,model=self.ai.generate_json("translate",TRANSLATION_PROMPT,piece,None,0.1); translated=compact_job_text(response.get("translation",""))
            if not translated: raise ValueError(f"Translation chunk {index} was empty")
            result.append(translated); models.append(model)
        translation="\n\n".join(result); model=", ".join(dict.fromkeys(models)); self.db.update_job(job_id,translation_zh=translation,translation_model=model,translation_hash=digest); return {"translation":translation,"model_used":model,"cached":False,"cancelled":False}
    @staticmethod
    def open_job(job: dict) -> None:
        url=str(job.get("url") or "")
        if not url.startswith(("http://","https://")): raise ValueError("Job URL is not valid")
        webbrowser.open(url)

class ResumeService:
    def __init__(self, db: Database, ai: AIManager): self.db,self.ai,self.paths=db,ai,get_paths()
    def import_original(self, source: str) -> Path:
        src=Path(source); suffix=src.suffix.casefold()
        if not src.is_file() or suffix not in SUPPORTED_RESUME_IMPORT_EXTENSIONS: raise ValueError("CareerOS requires an editable DOCX resume.\nPlease upload the original .docx file.")
        if src.stat().st_size > 25 * 1024 * 1024:
            raise ValueError("Resume is larger than the 25 MB safe processing limit.")
        try:
            with zipfile.ZipFile(src) as archive:
                members = archive.infolist(); expanded = sum(max(0, item.file_size) for item in members); compressed = sum(max(1, item.compress_size) for item in members)
                if len(members) > 5000 or expanded > 100 * 1024 * 1024 or expanded / compressed > 100:
                    raise ValueError("Resume archive expands beyond the safe processing limit.")
        except zipfile.BadZipFile as exc:
            raise ValueError("The selected file is not a valid DOCX document.") from exc
        stamp=datetime.now().strftime("%Y%m%d-%H%M%S-%f"); backup=self.paths.resumes_original/f"source-{stamp}{suffix}"; shutil.copy2(src,backup)
        try:
            document, blocks = parse_docx_resume(backup); report = ResumeIntegrityValidator().report(document); document.import_status = report["status"]; document.import_issues = report["issues"]; text = document.source_text
        except Exception as exc:
            backup.unlink(missing_ok=True); log.exception("DOCX import failed: %s", src); raise ValueError("CareerOS could not read this Word document.\nPlease open the resume in Microsoft Word or LibreOffice,\nsave it again as a standard .docx file, and retry.") from exc
        if not text.strip(): backup.unlink(missing_ok=True); raise ValueError("CareerOS could not read this Word document.\nPlease open the resume in Microsoft Word or LibreOffice,\nsave it again as a standard .docx file, and retry.")
        target=self.paths.resumes_original/f"original-{stamp}.txt"; target.write_text(text.strip(),encoding="utf-8")
        target.with_suffix(".json").write_text(json.dumps(resume_document_to_dict(document), ensure_ascii=False, indent=2), encoding="utf-8")
        target.with_suffix(".blocks.json").write_text(json.dumps(blocks_to_dict(blocks), ensure_ascii=False, indent=2), encoding="utf-8")
        # Scores belong to the resume that produced them.  A new import clears
        # AI results and recalculates only the deterministic preliminary score.
        self.db.clear_job_scores()
        for job in self.db.jobs():
            description = str(job.get("description") or "")
            if description:
                score = JobService._rule_score(text, str(job.get("title") or "") + "\n" + description, str(job.get("location") or ""))[0]
                self.db.update_job(int(job["id"]), rule_score=score)
        return target

    @staticmethod
    def parse_document(text: str) -> ResumeDocument: return parse_resume_text(text)

    @staticmethod
    def default_style() -> ResumeStyle:
        options = load_settings().get("resume_pdf", {}); spacing = {"Compact": 4.0, "Normal": 6.0, "Spacious": 9.0}.get(str(options.get("section_spacing", "Normal")), 6.0); return ResumeStyle(template_id=str(options.get("style", "Modern")).casefold(), font_family=str(options.get("font_family", "Helvetica")), body_font_size=float(options.get("font_size", 9)), line_spacing=float(options.get("line_spacing", 1.0)), section_spacing_before=spacing, section_spacing_after=max(2.0, spacing / 2.4))

    @staticmethod
    def default_layout() -> ResumeLayout:
        options = load_settings().get("resume_pdf", {}); margin = {"Narrow": .50, "Standard": .68, "Comfortable": .82}.get(str(options.get("margins", "Standard")), .68); return ResumeLayout(margin_top=margin, margin_bottom=margin, margin_left=margin, margin_right=margin, layout_mode="single_column", hidden_sections=["summary"])

    def document_from_version(self, version: dict) -> ResumeDocument:
        try:
            if str(version.get("document_json") or "").strip(): return resume_document_from_dict(json.loads(version["document_json"]))
        except (ValueError, TypeError, json.JSONDecodeError):
            log.warning("Resume version %s has malformed document_json; using legacy content", version.get("id"))
        return parse_resume_text(str(version.get("content") or ""))

    def style_from_version(self, version: dict) -> ResumeStyle:
        try: return resume_style_from_dict(json.loads(version.get("style_json") or "{}"))
        except (ValueError, TypeError, json.JSONDecodeError): return self.default_style()

    def layout_from_version(self, version: dict) -> ResumeLayout:
        try: return resume_layout_from_dict(json.loads(version.get("layout_json") or "{}"))
        except (ValueError, TypeError, json.JSONDecodeError): return self.default_layout()

    def update_version_document(self, version_id: int, document: ResumeDocument, style: ResumeStyle | None = None, layout: ResumeLayout | None = None) -> str:
        # Retire Summary on every editable path; legacy records are left intact
        # until a user explicitly saves them again.
        document.summary = ""
        ResumeIntegrityValidator().validate(document)
        effective_layout = layout or self.default_layout()
        if "summary" not in effective_layout.hidden_sections:
            effective_layout.hidden_sections.append("summary")
        content = resume_document_to_text(document, effective_layout.hidden_sections); self.db.update_resume_version(version_id, content=content, document_json=json.dumps(resume_document_to_dict(document)), style_json=json.dumps(resume_style_to_dict(style or self.default_style())), layout_json=json.dumps(resume_layout_to_dict(effective_layout))); return content

    def update_cv_letter(self, version_id: int, letter: str) -> str:
        letter = str(letter or "").strip()
        if not letter:
            raise ValueError("CV letter cannot be empty")
        version = next((value for value in self.db.resume_versions() if int(value.get("id") or -1) == int(version_id)), None)
        if not version or str(version.get("document_type", "")).upper() != "CV":
            raise ValueError("The selected draft is not a CV letter")
        if version.get("approved"):
            raise ValueError("Approved CV versions cannot be edited")
        metadata = self.cv_letter_metadata(version)
        prospective = {**version, "content": letter, "document_json": json.dumps(metadata), "template_version": "cv-letter-v1", "layout_version": "letter-v1"}
        self.render_version_outputs(prospective)
        self.db.update_resume_version(version_id, content=letter, document_json=json.dumps(metadata), template_version="cv-letter-v1", layout_version="letter-v1")
        return letter

    def render_version_outputs(self, version: dict) -> tuple[Path, Path]:
        style, layout = self.style_from_version(version), self.layout_from_version(version)
        name = version["version_name"]; pdf_path = self.paths.resumes_generated / f"{name}.pdf"; docx_path = self.paths.resumes_generated / f"{name}.docx"
        current_options = load_settings().get("resume_pdf", {})
        saved_options = {**current_options, "font_family": style.font_family, "font_size": style.body_font_size, "name_font_size": style.name_font_size, "section_font_size": style.section_font_size, "line_spacing": style.line_spacing, "paragraph_spacing": style.paragraph_spacing, "section_spacing_before": style.section_spacing_before, "section_spacing_after": style.section_spacing_after, "margin_top": layout.margin_top, "margin_bottom": layout.margin_bottom, "margin_left": layout.margin_left, "margin_right": layout.margin_right}
        if self.is_cv_letter_version(version):
            metadata = self.cv_letter_metadata(version)
            render_cv_letter_docx(version.get("content", ""), docx_path, metadata["candidate"], metadata["company"], metadata["title"], metadata["job_location"], metadata["date"], saved_options)
            if export_docx_pdf(docx_path, pdf_path):
                log.info("cv_letter_export version=%s pdf_backend=word-com docx=%s pdf=%s", version.get("id"), docx_path, pdf_path)
                return pdf_path, docx_path
            log.warning("cv_letter_export version=%s pdf_backend=reportlab-fallback docx=%s", version.get("id"), docx_path)
            render_cv_letter_pdf(version.get("content", ""), pdf_path, metadata["candidate"], metadata["company"], metadata["title"], metadata["job_location"], metadata["date"], saved_options)
            return pdf_path, docx_path

        document = self.document_from_version(version)
        if str(version.get("document_json") or "").strip():
            ResumeIntegrityValidator().validate(document)
        shared_options = saved_options
        is_cv = str(version.get("document_type", "Resume")).upper() == "CV"
        if is_cv:
            render_structured_cv_docx(document, docx_path, shared_options, layout.hidden_sections)
        else:
            render_structured_resume_docx(document, docx_path, shared_options, layout.hidden_sections)
        if export_docx_pdf(docx_path, pdf_path):
            log.info("%s_export version=%s pdf_backend=word-com docx=%s pdf=%s", "cv" if is_cv else "resume", version.get("id"), docx_path, pdf_path)
            return pdf_path, docx_path
        log.warning("%s_export version=%s pdf_backend=reportlab-fallback docx=%s", "cv" if is_cv else "resume", version.get("id"), docx_path)
        if is_cv:
            render_cv_document_pdf(document, pdf_path, style, layout)
        else:
            render_resume_document_pdf(document, pdf_path, style, layout)
        return pdf_path, docx_path
    def render_editor_preview(self, document: ResumeDocument, style: ResumeStyle, layout: ResumeLayout, version_id: int, generation: int, document_type: str = "Resume") -> Path:
        """Render the editor's working copy to a real PDF page.

        A unique filename avoids QPdfDocument/Windows locking conflicts.  On
        Windows the preview uses the same DOCX -> Word PDF path as a saved
        Professional draft, so pagination and typography match the final file.
        The ReportLab path is only a no-Word fallback.
        """
        ResumeIntegrityValidator().validate(document)
        stem = f"editor-preview-{int(version_id)}-{int(generation)}"
        docx_path = self.paths.cache / f"{stem}.docx"
        pdf_path = self.paths.cache / f"{stem}.pdf"
        options = load_settings().get("resume_pdf", {})
        shared_options = {
            **options,
            "font_family": style.font_family,
            "font_size": style.body_font_size,
            "name_font_size": style.name_font_size,
            "section_font_size": style.section_font_size,
            "line_spacing": style.line_spacing,
            "paragraph_spacing": style.paragraph_spacing,
            "section_spacing_before": style.section_spacing_before,
            "section_spacing_after": style.section_spacing_after,
            "margin_top": layout.margin_top,
            "margin_bottom": layout.margin_bottom,
            "margin_left": layout.margin_left,
            "margin_right": layout.margin_right,
        }
        is_cv = str(document_type).upper() == "CV"
        if is_cv:
            render_structured_cv_docx(document, docx_path, shared_options, layout.hidden_sections)
        else:
            render_structured_resume_docx(document, docx_path, shared_options, layout.hidden_sections)
        if not export_docx_pdf(docx_path, pdf_path):
            (render_cv_document_pdf if is_cv else render_resume_document_pdf)(document, pdf_path, style, layout)
        try:
            docx_path.unlink(missing_ok=True)
        except OSError:
            pass
        return pdf_path

    def original_path(self) -> Path | None:
        copies=sorted(self.paths.resumes_original.glob("original-*.txt"),key=lambda p:p.stat().st_mtime)
        if copies: return copies[-1]
        configured=Path(str(load_settings().get("resume_path","")).strip()) if str(load_settings().get("resume_path","")).strip() else None
        return configured if configured and configured.is_file() else None
    def original_text(self) -> str:
        path=self.original_path()
        if not path: raise FileNotFoundError("Import an original resume first")
        return path.read_text(encoding="utf-8",errors="replace")
    def original_document(self, source_path: Path | str | None = None) -> ResumeDocument:
        path = Path(source_path) if source_path else self.original_path()
        if not path: raise FileNotFoundError("Import an original resume first")
        if not path.is_file():
            raise FileNotFoundError("The original resume snapshot used by this draft is unavailable. Create a new draft from the active resume instead.")
        sidecar = path.with_suffix(".json")
        if sidecar.is_file():
            try: return resume_document_from_dict(json.loads(sidecar.read_text(encoding="utf-8")))
            except (ValueError, TypeError, json.JSONDecodeError): log.warning("Original resume sidecar is malformed: %s", sidecar)
        source = self.imported_source_path() if source_path is None else None
        if source and source.suffix.casefold() == ".docx":
            try: return parse_docx_resume(source)[0]
            except Exception: log.exception("Could not reparse original DOCX: %s", source)
        return self.parse_document(path.read_text(encoding="utf-8", errors="replace"))
    def imported_source_path(self) -> Path | None:
        # source-* filenames contain an import timestamp.  Do not sort by file
        # mtime here: import_original() uses shutil.copy2(), which intentionally
        # preserves the source document's original modification time.  A newly
        # imported DOCX can therefore have an older mtime than a legacy PDF and
        # be mistaken for the inactive source.  Filename order reflects the
        # actual CareerOS import order.
        files = sorted(self.paths.resumes_original.glob("source-*"), key=lambda p: p.name)
        return files[-1] if files else self.original_path()
    def prepare_original_preview(self) -> Path | None:
        """Prepare the original resume preview without touching Qt widgets.

        This is safe to run from a startup worker while the splash screen is
        visible.  The UI can then load the already-rendered PDF before the main
        window is shown, avoiding a Word-COM pause on the first Resume & CV
        navigation.
        """
        source = self.imported_source_path()
        if source and source.suffix.casefold() == ".docx":
            target = self.paths.cache / f"original-source-preview-{source.stem}.pdf"
            if target.is_file() or export_docx_pdf(source, target):
                return target
            # Word conversion is optional; fall through to the text renderer.
        if source and source.suffix.casefold() == ".pdf":
            return source
        try:
            content = self.original_text()
        except FileNotFoundError:
            return None
        return self.preview_pdf(content, "original-preview")

    def supporting_context(self, limit: int=24000) -> str:
        parts=[]; remaining=limit
        for record in reversed(self.db.supporting_documents()):
            path=Path(record.get("extracted_path") or "")
            if record.get("extraction_status")!="ready" or not path.exists() or remaining<=0: continue
            text=path.read_text(encoding="utf-8",errors="replace").strip()[:min(6000,remaining)]
            if text: parts.append(f"ADDITIONAL FILE: {record['original_name']}\n{text}"); remaining-=len(text)
        return "\n\n".join(parts)
    def candidate_context(self, include_resume: bool=True, resume_text: str | None = None) -> str:
        profile=load_settings().get("profile",{}); parts=[]
        if include_resume: parts.append("RESUME:\n"+(self.original_text() if resume_text is None else resume_text))
        confirmed="\n".join(f"{key.replace('_',' ').title()}: {value}" for key,value in profile.items() if str(value).strip())
        if confirmed: parts.append("VERIFIED CANDIDATE INFORMATION:\n"+confirmed)
        if extra:=self.supporting_context(): parts.append(extra)
        return "\n\n".join(parts)

    @staticmethod
    def is_cv_letter_version(version: dict) -> bool:
        if str(version.get("document_type", "Resume")).upper() != "CV":
            return False
        if str(version.get("template_version") or "").casefold() == "cv-letter-v1":
            return True
        try:
            value = json.loads(version.get("document_json") or "{}")
            return isinstance(value, dict) and value.get("kind") == "cv_letter"
        except (TypeError, ValueError, json.JSONDecodeError):
            return False

    def _candidate_letter_identity(self, base: ResumeDocument | None = None) -> dict[str, str]:
        base = base or self.original_document(); info = base.personal_info; profile = load_settings().get("profile", {})
        profile_name = " ".join(str(profile.get(key) or "").strip() for key in ("first_name", "last_name")).strip()
        address = ", ".join(str(profile.get(key) or "").strip() for key in ("address", "city", "province", "postal_code") if str(profile.get(key) or "").strip())
        return {
            "full_name": profile_name or info.full_name,
            "email": str(profile.get("email") or "").strip() or info.email,
            "phone": str(profile.get("phone") or "").strip() or info.phone,
            "location": address or info.location,
            "linkedin": str(profile.get("linkedin_url") or "").strip() or info.linkedin,
        }

    def cv_letter_metadata(self, version: dict, job: dict | None = None, base: ResumeDocument | None = None) -> dict:
        try:
            saved = json.loads(version.get("document_json") or "{}")
        except (TypeError, ValueError, json.JSONDecodeError):
            saved = {}
        if not isinstance(saved, dict):
            saved = {}
        job = job or self.db.job(int(version.get("job_id") or 0)) or {}
        candidate = saved.get("candidate") if isinstance(saved.get("candidate"), dict) else self._candidate_letter_identity(base)
        return {
            "kind": "cv_letter",
            "candidate": candidate,
            "company": str(saved.get("company") or job.get("company") or "").strip(),
            "title": str(saved.get("title") or job.get("title") or "").strip(),
            "job_location": str(saved.get("job_location") or job.get("location") or "").strip(),
            "date": str(saved.get("date") or datetime.now().strftime("%B %d, %Y")),
        }

    @staticmethod
    def _validated_cv_letter(result: object, candidate_source: str, complete_source: str, candidate_name: str) -> tuple[str, list[str]]:
        if not isinstance(result, dict):
            raise ValueError("AI did not return a CV letter")
        letter = str(result.get("letter") or "").strip()
        if not letter:
            raise ValueError("AI did not return a CV letter")
        words = re.findall(r"\b[\w'’-]+\b", letter)
        if len(words) < 120 or len(words) > 550:
            raise ValueError("AI returned a CV letter outside the supported 120-550 word range")
        paragraphs = [value.strip() for value in re.split(r"\n\s*\n", letter) if value.strip()]
        if len(paragraphs) < 4 or not paragraphs[0].casefold().startswith("dear "):
            raise ValueError("AI did not return a complete professional letter structure")
        if candidate_name and candidate_name.casefold() not in letter.casefold():
            raise ValueError("AI omitted the verified candidate signature")
        forbidden = {"education", "technical skills", "skills", "experience", "work experience", "projects", "engineering projects", "curriculum vitae", "resume"}
        lines = [line.strip().rstrip(":").casefold() for line in letter.splitlines() if line.strip()]
        if any(line in forbidden for line in lines) or sum(bool(re.match(r"^\s*[-*•]", line)) for line in letter.splitlines()) >= 2:
            raise ValueError("AI returned a resume-style document instead of a CV letter")
        source_numbers = set(re.findall(r"\b\d+(?:\.\d+)?%?\b", complete_source))
        added_numbers = set(re.findall(r"\b\d+(?:\.\d+)?%?\b", letter)) - source_numbers
        if added_numbers:
            raise ValueError("AI returned unsupported numbers in the CV letter")
        protected = {"solidworks", "catia", "creo", "inventor", "autocad", "ansys", "abaqus", "matlab", "simulink", "python", "c++", "java", "sql", "gd&t", "cfd", "fea", "six sigma", "pmp", "lean", "sap", "jira", "labview"}
        source_folded = complete_source.casefold()
        if any(term in letter.casefold() and term not in source_folded for term in protected):
            raise ValueError("AI returned unsupported skills or tools in the CV letter")
        facts = result.get("facts_used")
        if not isinstance(facts, list) or not facts:
            raise ValueError("AI did not identify the verified candidate facts used in the CV letter")
        clean_facts = []
        for value in facts:
            if isinstance(value, dict):
                value = next((value.get(key) for key in ("quote", "excerpt", "fact", "evidence", "source_text", "claim") if str(value.get(key) or "").strip()), "")
            text = str(value or "").strip()
            if text:
                clean_facts.append(text)
        if not clean_facts or any(not ResumeService._cv_fact_is_grounded(fact, candidate_source) for fact in clean_facts):
            raise ValueError("AI cited candidate facts that are not present in the resume or Settings")
        return letter, [str(value) for value in result.get("warnings", []) if str(value).strip()]

    @staticmethod
    def _cv_fact_is_grounded(fact: str, candidate_source: str) -> bool:
        """Accept traceable paraphrases without weakening unsupported-fact checks.

        Local models often return a faithful paraphrase in ``facts_used`` even
        when asked for an exact excerpt. Requiring byte-for-byte equality made
        valid CV letters fail. This matcher first accepts normalized substrings,
        then compares meaningful word roots so punctuation and small wording
        changes (design/designed, test/testing) remain traceable.
        """
        normalized_fact = re.sub(r"\s+", " ", str(fact or "")).strip().casefold()
        normalized_source = re.sub(r"\s+", " ", str(candidate_source or "")).strip().casefold()
        if not normalized_fact or not normalized_source:
            return False
        if normalized_fact in normalized_source:
            return True
        stop = {
            "a", "an", "and", "are", "as", "at", "be", "been", "being", "by", "candidate", "for", "from", "has", "have", "having", "in", "is", "it", "of", "on", "or", "that", "the", "their", "this", "to", "using", "was", "were", "with", "work", "worked", "experience",
        }

        def root(token: str) -> str:
            token = token.casefold().strip("-._/")
            if any(ch in token for ch in "+#&"):
                return token
            if len(token) > 5 and token.endswith("ies"):
                return token[:-3] + "y"
            for suffix in ("ments", "ment", "ation", "ations", "ing", "ied", "ed", "es", "s"):
                if len(token) > len(suffix) + 3 and token.endswith(suffix):
                    return token[:-len(suffix)]
            return token

        pattern = r"[a-z0-9][a-z0-9+#&./-]*"
        fact_tokens = [root(token) for token in re.findall(pattern, normalized_fact) if root(token) not in stop]
        source_tokens = {root(token) for token in re.findall(pattern, normalized_source) if root(token) not in stop}
        if not fact_tokens:
            return False
        matched = sum(token in source_tokens for token in fact_tokens)
        required = 1 if len(fact_tokens) == 1 else 2
        return matched >= required and matched / len(fact_tokens) >= 0.60

    def _generate_cv_letter(self, job: dict, base_document: ResumeDocument, resume_text: str, source_path: Path, selected_model: str | None, replace_version: dict | None) -> dict:
        candidate_source = self.candidate_context(True, resume_text)
        job_source = f"{job['title']} at {job['company']}\nLocation: {job.get('location') or ''}\n{job['description'][:12000]}"
        request = f"CANDIDATE INFORMATION:\n{candidate_source}\n\nJOB POSTING:\n{job_source}{generation_instruction('cv')}"
        result, model = self.ai.generate_json("cv_generation", CV_GENERATION_PROMPT, request, selected_model, 0.35)
        candidate = self._candidate_letter_identity(base_document)
        try:
            letter, warnings = self._validated_cv_letter(result, candidate_source, candidate_source + "\n\n" + job_source, candidate["full_name"])
        except ValueError as first_error:
            # Never save an unvalidated response. A constrained repair attempt
            # makes local models recover from paraphrased facts_used metadata
            # without relaxing the no-invention boundary.
            repair = request + "\n\nREPAIR REQUIRED:\nThe previous response was rejected: " + str(first_error) + "\nReturn corrected JSON. facts_used must be short exact excerpts copied from CANDIDATE INFORMATION. Do not add any candidate claim."
            result, model = self.ai.generate_json("cv_generation_repair", CV_GENERATION_PROMPT, repair, selected_model, 0.15)
            letter, warnings = self._validated_cv_letter(result, candidate_source, candidate_source + "\n\n" + job_source, candidate["full_name"])
        metadata = self.cv_letter_metadata({"job_id": job["id"]}, job, base_document)
        metadata["candidate"] = candidate
        number = len(self.db.resume_versions()) + 1
        name = replace_version["version_name"] if replace_version else f"cv_v{number:03d}"
        changes = [{"operation": "generate_cv_letter", "entity_type": "document", "entity_id": "cv-letter", "bullet_id": "", "original": "", "suggested": letter, "reason": "AI-generated job-targeted application letter", "risk": "LOW"}]
        style = self.default_style(); layout_model = self.default_layout()
        record = {"job_id": job["id"], "parent_version_id": replace_version.get("id") if replace_version else None, "generation_job_id": job["id"], "version_name": name, "source_path": str(source_path), "content": letter, "changes_json": json.dumps(changes), "document_type": "CV", "document_json": json.dumps(metadata), "style_json": json.dumps(resume_style_to_dict(style)), "layout_json": json.dumps(resume_layout_to_dict(layout_model)), "template_version": "cv-letter-v1", "layout_version": "letter-v1", "model_used": model}
        # Render before changing the database, so a Word/PDF failure cannot replace
        # a previously usable review draft with an incomplete record.
        self.render_version_outputs(record)
        if replace_version:
            version_id = int(replace_version["id"]); self.db.replace_resume_version(version_id, **record)
        else:
            version_id = self.db.add_resume_version(**record)
        return {**result, "letter": letter, "changes": changes, "rejected_operations": [], "generation_warnings": warnings, "version_id": version_id, "version_name": name, "document_type": "CV", "content": letter, "model_used": model}
    def optimize(self, job: dict, selected_model: str|None=None, document_type: str="Resume", replace_version: dict|None=None) -> dict:
        kind="CV" if str(document_type).upper()=="CV" else "Resume"
        source_path = Path(str(replace_version.get("source_path") or "")) if replace_version else self.original_path()
        if not source_path or not source_path.is_file():
            raise FileNotFoundError("The original resume snapshot for this draft is unavailable. Create a new draft from the active resume instead.")
        original = source_path.read_text(encoding="utf-8", errors="replace")
        base_document = self.original_document(source_path); report = ResumeIntegrityValidator().report(base_document)
        if report["status"] in {"REVIEW_REQUIRED", "FAIL"}:
            raise ValueError("Resume import needs review before generation. Re-import the DOCX or contact support with the local log file.")
        if kind == "CV":
            return self._generate_cv_letter(job, base_document, original, source_path, selected_model, replace_version)
        else:
            extra = self.candidate_context(False)
            goal = "Create a concise tailored resume for this role using only the allowed existing edit targets."
            generation_warnings = []
            allowed_targets = []
            for item in base_document.experience:
                allowed_targets.append(f"experience {item.id} -> bullets: {', '.join(b.id for b in item.bullets) or '(none)'}")
            for item in base_document.projects:
                allowed_targets.append(f"project {item.id} -> bullets: {', '.join(b.id for b in item.bullets) or '(none)'}")
            for item in base_document.skills:
                allowed_targets.append(f"skills {item.id}")
            target_lines = "\n- ".join(allowed_targets) if allowed_targets else "(no editable targets)"
            target_note = ("\n\nALLOWED EDIT TARGETS (use these exact ids only):\n- " + target_lines + "\nDo not create or edit a summary. Do not reference custom_sections, education, languages, personal_info, certifications, or any id not listed above." "\nIf none of the listed targets needs a safe rewrite, return \"operations\": [] instead of guessing an entity id.")
            request=f"CANONICAL RESUME JSON (do not rebuild it):\n{json.dumps(resume_document_to_dict(base_document), ensure_ascii=False)}\n\n{extra}\n\nJOB:\n{job['title']} at {job['company']}\n{job['description'][:12000]}\nDOCUMENT TARGET: {goal}{target_note}{generation_instruction('resume')}"
            result,model=self.ai.generate_json("resume_optimization", RESUME_OPTIMIZATION_PROMPT,request,selected_model,0.2)
            document, changes, rejected_operations = apply_resume_operations_best_effort(base_document, result.get("operations"))
            if rejected_operations:
                log.warning("Skipped %d invalid resume AI operation(s): %s", len(rejected_operations), [item["reason"] for item in rejected_operations])
            if result.get("operations") and not changes:
                reasons = "; ".join(dict.fromkeys(item["reason"] for item in rejected_operations))
                raise ValueError(f"AI returned no usable resume edits. {reasons}".strip())
            document.summary = ""
        # Summary is retired product data.  It may remain in legacy JSON until
        # that record is edited, but is never carried into a newly generated draft.
        document.summary = ""
        revised=resume_document_to_text(document, ["summary"])
        style=self.default_style(); layout_model=self.default_layout()
        number=len(self.db.resume_versions())+1
        name = f"{replace_version['version_name']}_r{number:03d}" if replace_version else f"{'cv' if kind=='CV' else 'resume'}_v{number:03d}"
        record={"job_id":job["id"],"parent_version_id":replace_version.get("id") if replace_version else None,"generation_job_id":job["id"],"version_name":name,"source_path":str(source_path),"content":revised,"changes_json":json.dumps(changes),"document_type":kind,"document_json":json.dumps(resume_document_to_dict(document)),"style_json":json.dumps(resume_style_to_dict(style)),"layout_json":json.dumps(resume_layout_to_dict(layout_model)),"template_version":"v2","layout_version":"v2","model_used":model}
        # Render before changing persistence. A Word/PDF failure must not replace
        # the last usable review draft with a record whose files do not exist.
        if replace_version:
            record["version_name"] = replace_version["version_name"]
        self.render_version_outputs(record)
        # Regenerate replaces the same unapproved row and avoids duplicate
        # Resume/CV entries for one job while preserving approved copies.
        if replace_version:
            version_id = int(replace_version["id"])
            self.db.replace_resume_version(version_id, **record)
        else:
            version_id=self.db.add_resume_version(**record)
        name = record["version_name"]
        return {**result,"changes":changes,"rejected_operations":rejected_operations,"generation_warnings":generation_warnings,"version_id":version_id,"version_name":name,"document_type":kind,"content":revised,"model_used":model}
    @staticmethod
    def diff(original: str, modified: str) -> str: return "\n".join(difflib.unified_diff(original.splitlines(),modified.splitlines(),fromfile="Original",tofile="Modified",lineterm=""))
    def decide(self, version_id: int, approved: bool) -> Path|None:
        version=next(v for v in self.db.resume_versions() if v["id"]==version_id)
        if not approved:
            self.db.set_resume_decision(version_id, False); return None
        if str(version.get("document_type", "Resume")).upper() == "CV" and not self.is_cv_letter_version(version):
            raise ValueError("This is a legacy resume-style CV draft. Regenerate it before approval so CareerOS creates the required letter format.")
        # Publish both files before marking the record approved. Applications
        # must never point at an approved version whose output copy failed.
        generated_pdf, generated_docx = self.render_version_outputs(version)
        target = self.paths.resumes_approved / generated_pdf.name; target_docx = self.paths.resumes_approved / generated_docx.name
        shutil.copy2(generated_pdf, target); shutil.copy2(generated_docx, target_docx)
        self.db.set_resume_decision(version_id, True)
        return target
    def preview_pdf(self, content: str, name: str="resume-preview") -> Path:
        options=load_settings().get("resume_pdf",{}); return render_resume_pdf(content,self.paths.cache/f"{safe_filename(name)}.pdf",options)
    def preview_docx(self, content: str, name: str="resume-preview", changes: list[dict]|None=None) -> Path:
        path=self.paths.cache/f"{safe_filename(name)}.docx"; render_professional_one_page(content, path, load_settings().get("resume_pdf", {})); return path
    def generated_docx_path(self, version_name: str, approved: bool=False) -> Path:
        return (self.paths.resumes_approved if approved else self.paths.resumes_generated)/f"{version_name}.docx"

class SupportingDocumentService:
    TEXT_EXTENSIONS={".txt",".md",".csv",".json",".yaml",".yml",".xml",".html",".htm",".rtf",".log",".ini",".py",".c",".cpp",".h"}
    MAX_FILE_BYTES = 25 * 1024 * 1024
    MAX_ARCHIVE_BYTES = 100 * 1024 * 1024
    MAX_EXTRACTED_CHARS = 300_000
    MAX_PDF_PAGES = 120
    def __init__(self,db:Database): self.db,self.paths=db,get_paths()
    @classmethod
    def _validate_archive(cls, source: Path) -> None:
        with zipfile.ZipFile(source) as archive:
            members = archive.infolist()
            if len(members) > 5000:
                raise ValueError("Document archive contains too many files")
            expanded = sum(max(0, member.file_size) for member in members)
            compressed = sum(max(1, member.compress_size) for member in members)
            if expanded > cls.MAX_ARCHIVE_BYTES or expanded / compressed > 100:
                raise ValueError("Document archive expands beyond the safe processing limit")
    @classmethod
    def _bounded(cls, value: str) -> str:
        return str(value or "")[:cls.MAX_EXTRACTED_CHARS]
    def _extract(self,source:Path)->tuple[str,str]:
        suffix=source.suffix.casefold()
        if suffix in self.TEXT_EXTENSIONS: return self._bounded(source.read_text(encoding="utf-8",errors="replace")).strip(),"ready"
        if suffix==".pdf":
            from pypdf import PdfReader
            reader = PdfReader(str(source), strict=False)
            if len(reader.pages) > self.MAX_PDF_PAGES: raise ValueError("PDF has too many pages")
            parts=[]; size=0
            for page in reader.pages:
                text=(page.extract_text() or "").strip(); parts.append(text); size += len(text)
                if size >= self.MAX_EXTRACTED_CHARS: break
            return self._bounded("\n\n".join(parts)).strip(),"ready"
        if suffix==".docx":
            from docx import Document
            self._validate_archive(source); doc=Document(str(source)); return self._bounded("\n".join(p.text.strip() for p in doc.paragraphs if p.text.strip())),"ready"
        if suffix==".xlsx":
            from openpyxl import load_workbook
            self._validate_archive(source); book=load_workbook(str(source),read_only=True,data_only=True)
            try:
                lines=[]; size=0
                for sheet in book.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        if not any(x is not None for x in row): continue
                        line=" | ".join(str(x) for x in row if x is not None); lines.append(line); size += len(line)
                        if size >= self.MAX_EXTRACTED_CHARS: break
                    if size >= self.MAX_EXTRACTED_CHARS: break
                return self._bounded("\n".join(lines)),"ready"
            finally: book.close()
        return "","stored_unreadable"
    def import_file(self,source:str)->dict:
        src=Path(source)
        if not src.is_file(): raise ValueError("Select a file")
        if src.stat().st_size > self.MAX_FILE_BYTES: raise ValueError("File is larger than the 25 MB safe processing limit")
        stamp=datetime.now().strftime("%Y%m%d-%H%M%S-%f"); saved=self.paths.supporting_documents/f"{stamp}-{safe_filename(src.stem)}{src.suffix.casefold()}"; shutil.copy2(src,saved)
        try: text,status=self._extract(saved)
        except Exception as exc: log.warning("Supporting document was stored but not parsed (%s): %s", src.name, exc); text,status="","stored_unreadable"
        extracted=None
        if text: target=self.paths.supporting_documents/f"{stamp}-{safe_filename(src.stem)}.extracted.txt"; target.write_text(text,encoding="utf-8"); extracted=str(target)
        elif status=="ready": status="stored_unreadable"
        ident=self.db.add_supporting_document(src.name,str(saved),extracted,status,len(text)); return {"id":ident,"name":src.name,"status":status,"characters":len(text)}
    def remove(self,document_id:int)->None:
        row=self.db.remove_supporting_document(document_id)
        if row:
            root = self.paths.supporting_documents.resolve()
            for value in (row.get("stored_path"),row.get("extracted_path")):
                if not value: continue
                path = Path(value).resolve()
                if path == root or root not in path.parents:
                    log.warning("Refused to delete supporting-document path outside storage: %s", path)
                    continue
                path.unlink(missing_ok=True)
